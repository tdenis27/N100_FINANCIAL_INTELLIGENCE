from pathlib import Path

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


ROOT = Path(__file__).resolve().parents[2]
OUTPUT_FILE = ROOT / "output" / "peer_comparison.xlsx"


GREEN = PatternFill(
    "solid",
    fgColor="C6EFCE"
)

YELLOW = PatternFill(
    "solid",
    fgColor="FFEB9C"
)

RED = PatternFill(
    "solid",
    fgColor="FFC7CE"
)

GOLD = PatternFill(
    "solid",
    fgColor="FFD966"
)


def export_peer_comparison(
    df,
    percentile_df
):

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    percentile_wide = percentile_df.pivot_table(
        index="company_id",
        columns="metric",
        values="percentile_rank",
        aggfunc="first"
    )

    percentile_wide.columns = [
        f"{column}_percentile"
        for column in percentile_wide.columns
    ]

    percentile_wide = (
        percentile_wide
        .reset_index()
    )

    report = df.merge(
        percentile_wide,
        on="company_id",
        how="left"
    )

    groups = [
        group
        for group in
        report["peer_group_name"]
        .dropna()
        .unique()
    ]

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        for group_name in groups:

            group = report[
                report["peer_group_name"]
                == group_name
            ].copy()

            numeric = group.select_dtypes(
                include="number"
            )

            median = numeric.median(
                numeric_only=True
            )

            summary = {
                column: ""
                for column in group.columns
            }

            summary["company_name"] = (
                "Peer Group Median"
            )

            for column in numeric.columns:
                summary[column] = median[column]

            group = pd.concat(
                [
                    group,
                    pd.DataFrame([summary])
                ],
                ignore_index=True
            )

            sheet = str(group_name)[:31]

            group.to_excel(
                writer,
                sheet_name=sheet,
                index=False
            )

    workbook = load_workbook(
        OUTPUT_FILE
    )

    for worksheet in workbook.worksheets:

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        percentile_columns = [
            column
            for name, column in headers.items()
            if name
            and str(name).endswith(
                "_percentile"
            )
        ]

        for row in range(
            2,
            worksheet.max_row + 1
        ):

            for column in percentile_columns:

                cell = worksheet.cell(
                    row=row,
                    column=column
                )

                try:
                    value = float(cell.value)
                except (TypeError, ValueError):
                    continue

                if value >= 0.75:
                    cell.fill = GREEN

                elif value <= 0.25:
                    cell.fill = RED

                else:
                    cell.fill = YELLOW

        last_row = worksheet.max_row

        for cell in worksheet[last_row]:
            cell.font = Font(bold=True)

    workbook.save(
        OUTPUT_FILE
    )

    print(
        f"Created: {OUTPUT_FILE}"
    )


if __name__ == "__main__":
    print(
        "Peer comparison report module ready."
    )